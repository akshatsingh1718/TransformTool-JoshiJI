import pandas as pd
from openpyxl.utils import get_column_letter
import copy
from datetime import datetime
import os
import json


class BaseTransformExcel():

    def read_config(cls, path: str):
        if os.path.exists(path):
            with open(path, "r") as f:
                return json.loads(f)
        return None


class TransformExcelPurchase(BaseTransformExcel):

    def __init__(cls, config: dict) -> None:
        cls.output_columns = config["output_columns"]
        cls.gst_data = config["gst_data"]
        cls.default_output_row = config["default_output_row"]
        cls.columns_to_sum = config["columns_to_sum"]
        cls.target_columns_index = config["target_columns_index"]
        cls.file_save_dir = "transformed"
        cls.perfix_for_totals = config["perfix_for_totals"]
        cls.party_gst_index = config["party_gst_index"]

    @staticmethod
    def default_config():
        return dict(
            output_columns=[
                "Inv No.",
                "Inv Date",
                "Party/Cash",
                "Product's Name",
                "HSN Code",
                "Qty",
                "Rate",
                "GST%",
                "Amount",
                "CGST %",
                "CGST Amt",
                "SGST %",
                "SGST Amt",
                "IGST Amt",
                "Discount %",
                "Total",
                "State",
                "Party's GST",
                "Reg Type",
                "Place of Supply",
                "Country",
                "Consignee State",
                "Consignee GST",
            ],
            columns_to_sum=["Amount", "CGST Amt", "SGST Amt", "IGST Amt", "Total"],
            perfix_for_totals=dict(column="Party/Cash", label="Total"),
            gst_data=[
                dict(gst_percentage=5, target_column_idx=11),
                dict(gst_percentage=12, target_column_idx=14),
                dict(gst_percentage=18, target_column_idx=17),
            ],
            target_columns_index=[
                dict(column="Inv Date", target_column_idx=0),
                dict(column="Inv No.", target_column_idx=1),
                dict(column="Party/Cash", target_column_idx=2),
                dict(column="Party's GST", target_column_idx=3),
            ],
            party_gst_index=3,
            default_output_row={
                "State": "Uttarakhand",
                "Place of Supply": "Uttarakhand",
                "Country": "India",
                "Consignee State": "Uttarakhand",
                "discount": 0.0,
            },
        )

    def get_default_row_format(cls, presets: dict) -> dict:
        row_dict: dict = cls.default_output_row

        for col in cls.output_columns:

            if col in presets.keys():
                row_dict[col] = presets[col]
                continue

            if col not in row_dict.keys():
                row_dict[col] = ""

        return row_dict

    def get_transformed_rows(cls, row):
        # Extract required values from the row
        # inv_date = row.iloc[0]
        # inv_no = row.iloc[1]
        # party_or_cash = row.iloc[2]

        # if start with 05 then calculate cgst/sgst
        # if not start with 05 then calculate igst
        gst_no = row.iloc[cls.party_gst_index]
        calculate_igst = not (str(gst_no).strip().startswith("05"))

        # amount_5 = float(row.iloc[11])
        # amount_12 = float(row.loc[14])
        # amount_18 = float(row.loc[17])

        new_rows = []
        # for per, amount in [(5, amount_5), (12, amount_12), (18, amount_18)]:
        for gst in cls.gst_data:
            cgst_sgst_per = (gst["gst_percentage"] / 2) if (not calculate_igst) else 0.0
            amount = row.iloc[gst["target_column_idx"]]

            cgst_sgst_amt = (
                amount * (cgst_sgst_per / 100) if not calculate_igst else 0.0
            )
            igst_amt = (amount * gst["gst_percentage"] / 100) if calculate_igst else 0.0
            discount = 0.0
            new_row = cls.get_default_row_format(
                {
                    **{
                        i["column"]: row.iloc[i["target_column_idx"]]
                        for i in cls.target_columns_index
                    },
                    "GST%": gst["gst_percentage"],
                    "Amount": amount,
                    "CGST %": cgst_sgst_per,
                    "CGST Amt": cgst_sgst_amt,
                    "SGST %": cgst_sgst_per,
                    "SGST Amt": cgst_sgst_amt,
                    "IGST Amt": igst_amt,
                    "Discount %": discount,
                    "Total": amount + (2 * cgst_sgst_amt) + igst_amt - discount,
                }
            )
            # new_row = {
            #     "Inv No.": inv_no,
            #     "Inv Date": inv_date,
            #     "Party/Cash": party_or_cash,
            #     "Product's Name": "",
            #     "HSN Code": "",
            #     "Qty": "",
            #     "Rate": "",
            #     "GST%": per,
            #     "Amount": amount,
            #     "CGST %": cgst_per,
            #     "CGST Amt": cgst_amt,
            #     "SGST %": sgst_per,
            #     "SGST Amt": sgst_amt,
            #     "IGST Amt": igst_amt,
            #     "Discount %": discount,
            #     "Total": amount + cgst_amt + sgst_amt + igst_amt - discount,
            #     "State": "Uttarakhand",
            #     "Party's GST": gst_no,
            #     "Reg Type": "",
            #     "Place of Supply": "Uttarakhand",
            #     "Country": "India",
            #     "Consignee State": "Uttarakhand",
            #     "Consignee GST": "",
            # }
            new_rows.append(copy.deepcopy(new_row))

        return new_rows

    def transform_bill(cls, df: pd.DataFrame, save=True):
        rows_to_df = []
        # Iterate over rows
        for _, row in df.iterrows():
            # print(str(row.iloc[0]), str(row.iloc[0]).startswith("A"))
            # If the row starts with a date (assuming the date is in the first cell)
            if not pd.notna(row.iloc[0]):
                continue

            # Check for date in 1st column
            try:
                pd.to_datetime(row.iloc[0], format="%d/%m/%Y").strftime(
                    format="%d/%m/%Y"
                )
            except Exception as e:
                continue

            # If the row starts with a bill number
            rows_to_df += cls.get_transformed_rows(row)

        # Get the current date and time
        current_datetime = datetime.now()
        # Format the date and time as desired
        formatted_datetime = current_datetime.strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"invoice_{formatted_datetime}.xlsx"
        xl_save_path = os.path.join(cls.file_save_dir, filename)

        # if dir not exist then create one
        if not os.path.isdir(cls.file_save_dir):
            os.makedirs(cls.file_save_dir, exist_ok=True)

        writer = pd.ExcelWriter(xl_save_path, engine="xlsxwriter")

        result_df = pd.DataFrame(rows_to_df, columns=cls.output_columns)

        # Convert the dataframe to an XlsxWriter Excel object.
        result_df.to_excel(writer, sheet_name="Sheet1", index=False)

        # Get the xlsxwriter objects from the dataframe writer object.
        workbook = writer.book
        worksheet = writer.sheets["Sheet1"]

        no_of_row = worksheet.dim_rowmax
        for col_to_sum in cls.columns_to_sum:
            col_letter = get_column_letter(
                result_df.columns.get_loc(col_to_sum) + 1
            )  # offset of 1 for index to pos

            index_of_sheet = f"{col_letter}{no_of_row + 2}"
            formula = f"=SUM({col_letter}2:{col_letter}{no_of_row + 1})"

            worksheet.write_formula(index_of_sheet, formula)

        if len(cls.columns_to_sum) > 0:
            col_letter = get_column_letter(
                result_df.columns.get_loc(cls.perfix_for_totals["column"]) + 1
            )  # offset of 1 for index to pos
            index_of_sheet = f"{col_letter}{no_of_row + 2}"
            worksheet.write_string(index_of_sheet, cls.perfix_for_totals["label"])

        if save:
            print(f"File saved to: {xl_save_path}")
            writer.close()
        return dict(
            xl_save_path = os.path.abspath(xl_save_path),
            save_dir = os.path.abspath(cls.file_save_dir)
        )

class TransformExcelSale(BaseTransformExcel):

    @staticmethod
    def default_config():
        return dict(
            output_columns=[
                "Inv No.",
                "Inv Date",
                "Party/Cash",
                "Product's Name",
                "HSN Code",
                "Qty",
                "Rate",
                "GST%",
                "Amount",
                "CGST %",
                "CGST Amt",
                "SGST %",
                "SGST Amt",
                "IGST Amt",
                "Discount %",
                "Total",
                "State",
                "Party's GST",
                "Reg Type",
                "Place of Supply",
                "Country",
                "Consignee State",
                "Consignee GST",
            ],
            columns_to_sum=["Amount", "CGST Amt", "SGST Amt", "IGST Amt", "Total"],
            perfix_for_totals=dict(column="Party/Cash", label="Total"),
            gst_data=[
                dict(gst_percentage=5, target_column_idx=4),
                dict(gst_percentage=12, target_column_idx=7),
                dict(gst_percentage=18, target_column_idx=10),
            ],
            target_columns_index=[
                dict(column="Inv No.", target_column_idx=0),
                dict(column="Party/Cash", target_column_idx=1),
            ],
            default_output_row={
                "State": "Uttarakhand",
                "Place of Supply": "Uttarakhand",
                "Country": "India",
                "Consignee State": "Uttarakhand",
                "discount": 0.0,
            },
        )

    def __init__(cls, config: dict) -> None:
        cls.output_columns = config["output_columns"]
        cls.gst_data = config["gst_data"]
        cls.default_output_row = config["default_output_row"]
        cls.columns_to_sum = config["columns_to_sum"]
        cls.target_columns_index = config["target_columns_index"]
        cls.file_save_dir = "transformed"
        cls.perfix_for_totals = config["perfix_for_totals"]

    def get_default_row_format(cls, presets: dict) -> dict:
        row_dict: dict = cls.default_output_row

        for col in cls.output_columns:

            if col in presets.keys():
                row_dict[col] = presets[col]
                continue

            if col not in row_dict.keys():
                row_dict[col] = ""

        return row_dict

    def get_transformed_rows(cls, row, bill_date, calculate_igst=False):
        new_rows = []
        for gst in cls.gst_data:
            cgst_sgst_per = gst["gst_percentage"] / 2
            amount = row.iloc[gst["target_column_idx"]]

            cgst_sgst_amt = amount * (cgst_sgst_per / 100)
            igst_amt = amount * (gst["gst_percentage"] / 100) if calculate_igst else 0.0
            discount = 0.0
            new_row = cls.get_default_row_format(
                {
                    **{
                        i["column"]: row.iloc[i["target_column_idx"]]
                        for i in cls.target_columns_index
                    },
                    "Inv Date": bill_date,
                    "GST%": gst["gst_percentage"],
                    "Amount": amount,
                    "CGST %": cgst_sgst_per,
                    "CGST Amt": cgst_sgst_amt,
                    "SGST %": cgst_sgst_per,
                    "SGST Amt": cgst_sgst_amt,
                    "IGST Amt": igst_amt,
                    "Discount %": discount,
                    "Total": amount + (2 * cgst_sgst_amt) + igst_amt - discount,
                }
            )
            new_rows.append(copy.deepcopy(new_row))

        return new_rows

    def transform_bill(cls, df: pd.DataFrame, save=True):
        rows_to_df = []
        # Iterate over rows
        for _, row in df.iterrows():
            # If the row starts with a date (assuming the date is in the first cell)
            if pd.notna(row.iloc[0]):
                try:
                    current_bill_date = pd.to_datetime(
                        row.iloc[0], format="%d/%m/%Y"
                    ).strftime(format="%d/%m/%Y")
                    continue
                except ValueError:
                    # If the value is not a valid date, move to the next row
                    pass

            # If the row starts with a bill number
            if (
                str(row.iloc[0]).startswith("A")
                and len(str(row.iloc[0]).split(" ")) == 1
                and len(str(row.iloc[0])) >= 6
            ):
                # Set current bill number
                rows_to_df += cls.get_transformed_rows(row, bill_date=current_bill_date)

        # Get the current date and time
        current_datetime = datetime.now()
        # Format the date and time as desired
        formatted_datetime = current_datetime.strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"invoice_{formatted_datetime}.xlsx"
        xl_save_path = os.path.join(cls.file_save_dir, filename)

        # if dir not exist then create one
        if not os.path.isdir(cls.file_save_dir):
            os.makedirs(cls.file_save_dir, exist_ok=True)

        writer = pd.ExcelWriter(xl_save_path, engine="xlsxwriter")

        result_df = pd.DataFrame(rows_to_df, columns=cls.output_columns)

        # Convert the dataframe to an XlsxWriter Excel object.
        result_df.to_excel(writer, sheet_name="Sheet1", index=False)

        # Get the xlsxwriter objects from the dataframe writer object.
        workbook = writer.book
        worksheet = writer.sheets["Sheet1"]

        no_of_row = worksheet.dim_rowmax
        for col_to_sum in cls.columns_to_sum:
            col_letter = get_column_letter(
                result_df.columns.get_loc(col_to_sum) + 1
            )  # offset of 1 for index to pos

            index_of_sheet = f"{col_letter}{no_of_row + 2}"
            formula = f"=SUM({col_letter}2:{col_letter}{no_of_row + 1})"

            worksheet.write_formula(index_of_sheet, formula)

        if len(cls.columns_to_sum) > 0:
            col_letter = get_column_letter(
                result_df.columns.get_loc(cls.perfix_for_totals["column"]) + 1
            )  # offset of 1 for index to pos
            index_of_sheet = f"{col_letter}{no_of_row + 2}"
            worksheet.write_string(index_of_sheet, cls.perfix_for_totals["label"])

        if save:
            print(f"File saved to: {xl_save_path}")
            writer.close()

        return dict(
            xl_save_path = os.path.abspath(xl_save_path),
            save_dir = os.path.abspath(cls.file_save_dir)
        )


def check_for_purchase():
    default_config = TransformExcelPurchase.default_config()

    obj = TransformExcelPurchase(default_config)

    def get_df():
        sale_path = "/home/akshat/Documents/projects/joshi-uncle/data/REPORT_2222.XLS"
        df = pd.read_excel(
            sale_path, sheet_name="MARG ERP 9+ Excel Report", header=None
        )
        return df

    obj.transform_bill(get_df())


def check_for_sale():
    default_config = TransformExcelSale.default_config()

    obj = TransformExcelSale(default_config)

    def get_df():
        sale_path = "data/SALE JAN FINAL 2024.XLS"
        df = pd.read_excel(
            sale_path, sheet_name="MARG ERP 9+ Excel Report", header=None
        )
        return df

    obj.transform_bill(get_df())


if __name__ == "__main__":
    check_for_purchase()
