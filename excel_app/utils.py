import pandas as pd
from openpyxl.utils import get_column_letter
import copy
from datetime import datetime
import os
import json
import math
from typing import List, Union

import warnings

warnings.simplefilter("ignore")


class BaseTransformExcel:

    @staticmethod
    def read_config(path: str):
        if os.path.exists(path):
            with open(path, "r") as f:
                return json.load(f)
        return None

    @staticmethod
    def get_filename_with_datetime(prefix="", suffix="", ext="xlsx"):
        # Get the current date and time
        current_datetime = datetime.now()
        # Format the date and time as desired
        formatted_datetime = current_datetime.strftime("%Y-%m-%d_%H-%M-%S")
        return f"{prefix}{formatted_datetime}{suffix}.{ext}"

    def get_default_row_format(cls, presets: dict) -> dict:

        # Here instead of default value prv row values is present
        row_dict: dict = copy.deepcopy(cls._default_output_row)

        # Iterate over output columns
        for col in cls.output_columns:

            # set preset values defined overriding default values
            if col in presets.keys():
                row_dict[col] = presets[col]
                continue

            if col not in row_dict.keys():
                row_dict[col] = ""

        return row_dict

    def get(cls, row, column: str):
        return row[cls.column_to_idx[column]]

    def fget(cls, row, column: str = "", dcol: str = ""):
        """
        Float get
        dcol: direct column name
        column: fetch index from column_to_idx
        """
        if dcol != "":
            return float(row[dcol])

        return row[cls.column_to_idx[column]]

    def post_processing(
        cls, df: pd.DataFrame, save=True, name_prefix: str = None, *args, **kwargs
    ):
        re_arrange_cols = kwargs.get("re_arrange_cols", False)

        if save and name_prefix is None:
            raise Exception("Please provide name prefix for xl file")

        filename = cls.get_filename_with_datetime(prefix=name_prefix)
        xl_save_path = os.path.join(cls.file_save_dir, filename)

        # if dir not exist then create one
        if save and (not os.path.isdir(cls.file_save_dir)):
            os.makedirs(cls.file_save_dir, exist_ok=True)

        if re_arrange_cols and cls.output_columns:
            df = df.reindex(cls.output_columns, axis=1)

        writer = pd.ExcelWriter(xl_save_path, engine="xlsxwriter")

        # Convert the dataframe to an XlsxWriter Excel object.
        df.to_excel(writer, sheet_name="Sheet1", index=False)

        # Get the xlsxwriter objects from the dataframe writer object.
        worksheet = writer.sheets["Sheet1"]

        if len(cls.columns_to_sum) > 0:
            no_of_row = worksheet.dim_rowmax
            for col_to_sum in cls.columns_to_sum:
                col_letter = get_column_letter(
                    df.columns.get_loc(col_to_sum) + 1
                )  # offset of 1 for index to pos

                index_of_sheet = f"{col_letter}{no_of_row + 2}"
                formula = f"=SUM({col_letter}2:{col_letter}{no_of_row + 1})"

                worksheet.write_formula(index_of_sheet, formula)

            col_letter = get_column_letter(
                df.columns.get_loc(cls.perfix_for_totals["column"]) + 1
            )  # offset of 1 for index to pos
            index_of_sheet = f"{col_letter}{no_of_row + 2}"
            worksheet.write_string(index_of_sheet, cls.perfix_for_totals["label"])

        if save:
            print(f"File saved to: {xl_save_path}")
            writer.close()

        return dict(
            xl_save_path=xl_save_path,
            save_dir=cls.file_save_dir,
            xl_file_name=filename,
            df=df,
        )

    @staticmethod
    def strip_column_names(df):
        """
        Strips leading and trailing whitespaces from column names in a DataFrame.

        Parameters:
        df (DataFrame): The pandas DataFrame.

        Returns:
        DataFrame: The DataFrame with stripped column names.
        """
        df.columns = [col.strip() for col in df.columns]
        return df


class EchsDueTransfromation(BaseTransformExcel):
    NAME = "ECHS DUE"
    CONFIG = "echs_due-config.json"

    def __init__(cls, config):
        cls.file_save_dir = config.get("file_save_dir")
        cls.output_columns = config.get("output_columns")
        cls.sheetname = config.get("sheetname")
        cls.file_prefix = config.get("file_prefix")
        cls.save_sheet_name = config.get("save_sheet_name")

    def transform(cls, file_paths: List, save=True) -> pd.DataFrame:
        # Read each file, skipping the first 7 rows
        read_df = lambda path: pd.read_excel(
            path, sheet_name=cls.sheetname, skiprows=2, skipfooter=1, header=None
        )

        df_list = []
        for path in file_paths:
            df = read_df(path)
            df["Date"] = "".join(os.path.basename(str(path)).split(".")[:-1])
            df_list.append(df)

        df = pd.concat(df_list, axis=0)
        df.columns = cls.output_columns

        filename = cls.get_filename_with_datetime(prefix=cls.file_prefix)
        xl_save_path = os.path.join(cls.file_save_dir, filename)

        # if dir not exist then create one
        if not os.path.isdir(cls.file_save_dir):
            os.makedirs(cls.file_save_dir, exist_ok=True)
        if save:
            print(f"File saved to: {xl_save_path}")
            df.to_excel(xl_save_path, sheet_name=cls.save_sheet_name, index=False)

        return dict(
            xl_save_path=xl_save_path, save_dir=cls.file_save_dir, xl_file_name=filename
        )


class GST2BTransfromation(BaseTransformExcel):
    NAME = "GST 2B Excel"
    CONFIG = "gst2bexcel_config.json"

    def __init__(cls, config):
        cls.file_save_dir = config.get("file_save_dir")
        cls.output_columns = config.get("output_columns")
        cls.sheetname = config.get("sheetname")
        cls.file_prefix = config.get("file_prefix")
        cls.save_sheet_name = config.get("save_sheet_name")

    def transform(cls, file_paths: List, save=True) -> pd.DataFrame:
        # Read each file, skipping the first 7 rows
        read_df = lambda path: pd.read_excel(path, sheet_name=cls.sheetname, skiprows=7)

        df_list = []
        for path in file_paths:
            for _, row in read_df(path).iterrows():
                df_list.append(list(row))

        df = pd.DataFrame(data=df_list, columns=cls.output_columns)
        df = df.sort_values(by="Trade/Legal name")

        filename = cls.get_filename_with_datetime(prefix=cls.file_prefix)
        xl_save_path = os.path.join(cls.file_save_dir, filename)

        # if dir not exist then create one
        if not os.path.isdir(cls.file_save_dir):
            os.makedirs(cls.file_save_dir, exist_ok=True)
        if save:
            # df.to_excel(xl_save_path, sheet_name=cls.save_sheet_name, index=False)
            print(f"File saved to: {xl_save_path}")
            df.to_excel(xl_save_path, sheet_name="Sheet1", index=False)

        return dict(
            xl_save_path=xl_save_path, save_dir=cls.file_save_dir, xl_file_name=filename
        )


class IpdOpdTransfromation(BaseTransformExcel):

    def __init__(cls, sort_columns: str, _for: str, filename_prefix=""):
        cls.sort_columns = sort_columns
        cls.filename_prefix = filename_prefix
        cls.file_save_dir = "transformed"
        cls._for = _for

    def transform(cls, path: str, save=True) -> pd.DataFrame:
        df = pd.read_excel(path)

        df = cls.strip_column_names(df)

        df = df.sort_values(by=cls.sort_columns)

        ## change date time fileds format
        # List to store the names of datetime columns
        datetime_columns = []
        # Identify datetime columns
        for column in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[column]):
                datetime_columns.append(column)
        # Convert datetime columns to date format="%d/%m/%Y"
        for column in datetime_columns:
            df[column] = df[column].dt.strftime("%d/%m/%Y")

        def concat_cols(r):
            if cls._for == "ipd":
                try:
                    return (
                        f"Bill No: {int(r['Bill No'])} {r['Patient Name']} {r['TPA']}"
                    )
                except Exception as e:
                    return ""
            try:
                return f"Bill No: {int(r['Bill No'])} {r['Patient Name']}"
            except Exception as e:
                return ""

        # concat columns
        df["Narration"] = df.apply(concat_cols, axis=1)

        # drop all the rows where all the cell are empty
        df.dropna(how="all", inplace=True)

        filename = cls.get_filename_with_datetime(prefix=f"{cls.filename_prefix}_")
        xl_save_path = os.path.join(cls.file_save_dir, filename)

        # if dir not exist then create one
        if not os.path.isdir(cls.file_save_dir):
            os.makedirs(cls.file_save_dir, exist_ok=True)

        if save:
            print(f"File saved to: {xl_save_path}")
            # Convert the dataframe to an XlsxWriter Excel object.
            df.to_excel(xl_save_path, sheet_name="Sheet1", index=False)
        return dict(
            xl_save_path=xl_save_path, save_dir=cls.file_save_dir, xl_file_name=filename
        )


class TransformExcelPurchase(BaseTransformExcel):

    def __init__(cls, config: dict) -> None:
        cls.output_columns = config["output_columns"]
        cls.gst_data = config["gst_data"]
        cls._default_output_row = config["default_output_row"]
        cls.columns_to_sum = config["columns_to_sum"]
        cls.target_columns_index = config["target_columns_index"]

        # constants
        cls.file_save_dir = "transformed"
        cls.filename_prefix = "GSTR2"

        cls.perfix_for_totals = config["perfix_for_totals"]
        cls.party_gst_index = config["party_gst_index"]
        cls.gst_mapping_xl_path: str = config["gst_mapping_xl_path"]
        cls.column_to_idx: dict = config["column_to_idx"]

    def get_gst_state(cls, gst_code: str) -> str:
        try:
            # Get the absolute path to the Excel file
            excel_file_path = os.path.join(cls.gst_mapping_xl_path)

            # Read the Excel file into a DataFrame
            df = pd.read_excel(excel_file_path)

            # Check if the input_value exists in the 'Mapping from' column
            mapping_to = df.loc[df["TIN"] == gst_code, "State"].values

            if len(mapping_to) > 0:
                # Return the corresponding mapping to value
                return mapping_to[0]
            else:
                # If input_value not found, return None
                return None
        except Exception as e:
            print(f"An error occurred: {e}")
            return None

    def get_transformed_rows(cls, row):
        # if start with 05 then calculate cgst/sgst
        # if not start with 05 then calculate igst
        gst_no = str(row.iloc[cls.party_gst_index]).strip()
        calculate_igst = not (str(gst_no).strip().startswith("05"))
        state = "" if gst_no == "nan" else cls.get_gst_state(int(gst_no[:2]))

        new_rows = []
        # for per, amount in [(5, amount_5), (12, amount_12), (18, amount_18)]:
        for gst in cls.gst_data:
            gst_percentage = gst["gst_percentage"]
            cgst_sgst_per = (gst_percentage / 2) if (not calculate_igst) else 0.0
            amount = float(row.iloc[gst["target_column_idx"]])

            if amount == 0.0:
                continue

            cgst_sgst_amt = (
                round(amount * cgst_sgst_per / 100, 2) if not calculate_igst else 0.0
            )
            igst_amt = (
                round(amount * gst_percentage / 100, 2) if calculate_igst else 0.0
            )
            discount = 0.0
            total_amt = round(amount + (2 * cgst_sgst_amt) + igst_amt - discount, 2)
            new_row = cls.get_default_row_format(
                {
                    **{
                        i["column"]: str(row.iloc[i["target_column_idx"]]).strip()
                        for i in cls.target_columns_index
                    },
                    "Party's GST": "" if gst_no == "nan" else gst_no,
                    "Reg Type": (
                        "unregistered/consumer" if gst_no == "nan" else "regular"
                    ),
                    "Product's Name": f"Medicine {gst_percentage}%",
                    "State": "Uttrakhand" if gst_no == "nan" else state,
                    "Party/Cash": str(row.iloc[cls.column_to_idx["Party/Cash"]])[
                        5:
                    ].strip(),
                    "GST%": gst_percentage,
                    "Amount": amount,
                    "CGST %": cgst_sgst_per,
                    "CGST Amt": cgst_sgst_amt,
                    "SGST %": cgst_sgst_per,
                    "SGST Amt": cgst_sgst_amt,
                    "IGST Amt": igst_amt,
                    "Discount %": discount,
                    "Total": total_amt,
                }
            )

            new_rows.append(copy.deepcopy(new_row))

        return new_rows

    def transform(cls, path: str, save=True):
        df = pd.read_excel(path)

        rows_to_df = []
        # Iterate over rows
        for _, row in df.iterrows():
            if not pd.notna(row.iloc[0]):
                continue

            # Check for date in 1st column
            try:
                pd.to_datetime(row.iloc[0], format="%d/%m/%Y").strftime(
                    format="%d/%m/%Y"
                )
            except Exception as e:
                continue

            # If the row starts with a bill number
            rows_to_df += cls.get_transformed_rows(row)

        filename = cls.get_filename_with_datetime(prefix=f"{cls.filename_prefix}_")
        xl_save_path = os.path.join(cls.file_save_dir, filename)

        # if dir not exist then create one
        if not os.path.isdir(cls.file_save_dir):
            os.makedirs(cls.file_save_dir, exist_ok=True)

        writer = pd.ExcelWriter(xl_save_path, engine="xlsxwriter")

        result_df = pd.DataFrame(rows_to_df, columns=cls.output_columns)

        # Convert the dataframe to an XlsxWriter Excel object.
        result_df.to_excel(writer, sheet_name="Sheet1", index=False)

        # Get the xlsxwriter objects from the dataframe writer object.
        workbook = writer.book
        worksheet = writer.sheets["Sheet1"]

        no_of_row = worksheet.dim_rowmax
        for col_to_sum in cls.columns_to_sum:
            col_letter = get_column_letter(
                result_df.columns.get_loc(col_to_sum) + 1
            )  # offset of 1 for index to pos

            index_of_sheet = f"{col_letter}{no_of_row + 2}"
            formula = f"=SUM({col_letter}2:{col_letter}{no_of_row + 1})"

            worksheet.write_formula(index_of_sheet, formula)

        if len(cls.columns_to_sum) > 0:
            col_letter = get_column_letter(
                result_df.columns.get_loc(cls.perfix_for_totals["column"]) + 1
            )  # offset of 1 for index to pos
            index_of_sheet = f"{col_letter}{no_of_row + 2}"
            worksheet.write_string(index_of_sheet, cls.perfix_for_totals["label"])

        if save:
            print(f"File saved to: {xl_save_path}")
            writer.close()
        return dict(
            xl_save_path=xl_save_path, save_dir=cls.file_save_dir, xl_file_name=filename
        )


class TransformExcelSale(BaseTransformExcel):

    def __init__(cls, config: dict, *args, **kwargs) -> None:
        cls.output_columns = config["output_columns"]
        cls.gst_data = config["gst_data"]
        cls._default_output_row = config["default_output_row"]
        cls.columns_to_sum = config["columns_to_sum"]
        cls.target_columns_index = config["target_columns_index"]
        cls.file_save_dir = "transformed"
        cls.filename_prefix = "GSTR1"
        cls.perfix_for_totals = config["perfix_for_totals"]
        cls.party_gst_index = config["party_gst_index"]
        cls.column_to_idx: dict = config["column_to_idx"]

        # changes every time
        cls.bill_no_prefix = kwargs.get("bill_no_prefix", None)
        cls.bill_no_suffix_counter = kwargs.get("bill_no_suffix_counter", None)
        cls.calculate_igst = kwargs.get("calculate_igst", False)

    def get_transformed_rows(
        cls, row, bill_date: str, bill_no: str, calculate_igst=False
    ):
        new_rows = []
        gst_no = str(row.iloc[cls.party_gst_index])

        for gst in cls.gst_data:
            gst_percentage = gst["gst_percentage"]
            cgst_sgst_per = gst_percentage / 2
            amount = float(row.iloc[gst["target_column_idx"]])
            if amount == 0.0:
                continue

            cgst_sgst_amt = round(amount * cgst_sgst_per / 100, 2)
            igst_amt = (
                round(amount * gst_percentage / 100, 2) if calculate_igst else 0.0
            )
            discount = 0.0
            total_amt = round(amount + (2 * cgst_sgst_amt) + igst_amt - discount, 2)
            new_row = cls.get_default_row_format(
                {
                    **{
                        i["column"]: str(row.iloc[i["target_column_idx"]]).strip()
                        for i in cls.target_columns_index
                    },
                    "Reg Type": (
                        "unregistered/consumer" if gst_no == "nan" else "regular"
                    ),
                    "Product's Name": f"Medicine {gst_percentage}%",
                    "Party/Cash": str(row.iloc[cls.column_to_idx["Party/Cash"]])
                    .split("-")[0]
                    .strip(),
                    "Bill No.": bill_no,
                    "Inv Date": bill_date,
                    "GST%": gst_percentage,
                    "Amount": amount,
                    "CGST %": cgst_sgst_per,
                    "CGST Amt": cgst_sgst_amt,
                    "SGST %": cgst_sgst_per,
                    "SGST Amt": cgst_sgst_amt,
                    "IGST Amt": igst_amt,
                    "Discount %": discount,
                    "Total": total_amt,
                }
            )
            new_rows.append(copy.deepcopy(new_row))

        return new_rows

    def transform(cls, path: str, save=True):
        df = pd.read_excel(path)
        rows_to_df = []
        bill_counter = cls.bill_no_suffix_counter

        # Iterate over rows
        for _, row in df.iterrows():
            # If the row starts with a date (assuming the date is in the first cell)
            if pd.notna(row.iloc[0]):
                try:
                    current_bill_date = pd.to_datetime(
                        row.iloc[0], format="%d/%m/%Y"
                    ).strftime(format="%d/%m/%Y")
                    continue
                except ValueError:
                    # If the value is not a valid date, move to the next row
                    pass

            # If the row starts with a bill number
            if (
                str(row.iloc[0]).startswith("A")
                and len(str(row.iloc[0]).split(" ")) == 1
                and len(str(row.iloc[0])) >= 6
            ):
                # Set current bill number
                new_tranformed_rows = cls.get_transformed_rows(
                    row,
                    bill_date=current_bill_date,
                    bill_no=f"{cls.bill_no_prefix}{bill_counter:05d}",
                    calculate_igst=cls.calculate_igst,
                )
                rows_to_df += new_tranformed_rows

                # if new rows added then only increase the bill counter
                # there may be case where no rows were added because amount is 0
                if len(new_tranformed_rows) > 0:
                    bill_counter += 1

        filename = cls.get_filename_with_datetime(prefix=f"{cls.filename_prefix}_")
        xl_save_path = os.path.join(cls.file_save_dir, filename)

        # if dir not exist then create one
        if not os.path.isdir(cls.file_save_dir):
            os.makedirs(cls.file_save_dir, exist_ok=True)

        writer = pd.ExcelWriter(xl_save_path, engine="xlsxwriter")

        result_df = pd.DataFrame(rows_to_df, columns=cls.output_columns)

        # Convert the dataframe to an XlsxWriter Excel object.
        result_df.to_excel(writer, sheet_name="Sheet1", index=False)

        # Get the xlsxwriter objects from the dataframe writer object.
        workbook = writer.book
        worksheet = writer.sheets["Sheet1"]

        no_of_row = worksheet.dim_rowmax
        for col_to_sum in cls.columns_to_sum:
            col_letter = get_column_letter(
                result_df.columns.get_loc(col_to_sum) + 1
            )  # offset of 1 for index to pos

            index_of_sheet = f"{col_letter}{no_of_row + 2}"
            formula = f"=SUM({col_letter}2:{col_letter}{no_of_row + 1})"

            worksheet.write_formula(index_of_sheet, formula)

        if len(cls.columns_to_sum) > 0:
            col_letter = get_column_letter(
                result_df.columns.get_loc(cls.perfix_for_totals["column"]) + 1
            )  # offset of 1 for index to pos
            index_of_sheet = f"{col_letter}{no_of_row + 2}"
            worksheet.write_string(index_of_sheet, cls.perfix_for_totals["label"])

        if save:
            print(f"File saved to: {xl_save_path}")
            writer.close()

        return dict(
            xl_save_path=xl_save_path, save_dir=cls.file_save_dir, xl_file_name=filename
        )


# Stock Transformation earlier
class TransformStockExcel(BaseTransformExcel):
    APP_NAME = "Mutual Fund"

    def __init__(cls, config: dict, *args, **kwargs) -> None:
        cls.file_save_dir = "transformed"

        pass

    def transform(cls, path: str, save=True) -> pd.DataFrame:
        """
        - Iterate over df rows
        - If row == header then append opening balalnce data.
        """
        df = pd.read_excel(path, skiprows=4)
        result = []
        stock_name = None
        for index, row in df.iterrows():
            # if the header found
            if (
                row[0] == "Sr."
                and row[1] == "Transaction Date"
                and row[2] == "Transaction Type"
                and row[3] == "Amount"
                and row[4] == "Units"
            ):
                # if header found then get stock name from prv cell
                stock_name = str(df.iloc[index - 1, 0]).strip().split("-")
                stock_name = (
                    "".join(stock_name[:-1]) if len(stock_name) > 1 else stock_name[0]
                )
                # if stock_name == "":
                #     [print("==============================")]
                #     print(f"Stock found at {index - 1}")
                #     print(str(df.iloc[index - 1, 0]).split("-")[:-1])
                # also append the next cell data to the result which is opening balance
                result.append(
                    [
                        len(result) + 1,
                        stock_name,
                        "",
                        df.iloc[index + 1, 1],
                        df.iloc[index + 1, 3],
                        "",
                    ]
                )
            # if the cell is not null and has int data at first
            elif (
                pd.notnull(row[0]) and isinstance(row[0], int) and str(row[2]) != "nan"
            ):

                result.append(
                    [len(result) + 1, stock_name, row[1], row[2], row[3], row[4]]
                )
                # if stock_name == "":
                #     print("find")
            # if nothing matches then continue
            else:
                continue

        columns = [
            "Sr.",
            "Fund Name",
            "Transaction Date",
            "Transaction Type",
            "Amount",
            "Units",
        ]

        df = pd.DataFrame(result, columns=columns)

        if save:
            filename = cls.get_filename_with_datetime(prefix="Mutual-Fund_")
            xl_save_path = os.path.join(cls.file_save_dir, filename)
            df.to_excel(xl_save_path, index=False)
            print(f"File saved to: {xl_save_path}")

        return dict(
            xl_save_path=xl_save_path, save_dir=cls.file_save_dir, xl_file_name=filename
        )


class TransformExcelGST(BaseTransformExcel):
    NAME = "GSTR1 With Qty"
    CONFIG = "gstWithQty_config.json"

    def __init__(cls, config: dict, *args, **kwargs) -> None:
        cls.file_save_dir = "transformed"
        cls.output_columns = config["output_columns"]
        cls._default_output_row = config["default_output_row"]
        cls.columns_to_sum = config["columns_to_sum"]
        cls.perfix_for_totals = config["perfix_for_totals"]
        cls.column_to_idx: dict = config["column_to_idx"]
        cls.target_columns_index: list = config["target_columns_index"]
        cls.xl_name_prefix: str = config["xl_name_prefix"]

        # changes every time
        cls.bill_no_prefix = kwargs.get("bill_no_prefix", None)
        cls.bill_no_suffix_counter = kwargs.get("bill_no_suffix_counter", None)
        cls.calculate_igst = kwargs.get("calculate_igst", False)

        # Mapping file
        cls._mapping_df = kwargs.get("mapping_df", None)  # optional
        cls.mapping_sheetname = config["mapping_sheetname"]

    def transform_mapping(cls):
        if cls._mapping_df is None:
            return

        df = cls._mapping_df.iloc[6:, 1:]
        # Set the headers to be the values from the first row
        new_headers = df.iloc[0]
        df.columns = new_headers

        # Reset the index
        df.reset_index(drop=True, inplace=True)

        cls._mapping_df = df

    def is_present_in_mapping(cls, value_to_find, column: str = "Bill No."):
        if cls._mapping_df is None:
            return False

        if value_to_find in cls._mapping_df[column].values:
            index = cls._mapping_df[cls._mapping_df[column] == value_to_find].index[0]
            if float(cls._mapping_df.at[index, "INDIAN BANK"]) > 0.0:
                return True
        return False

    def get_transformed_rows(cls, row, bill_no: str, *args, **kwargs):

        roundoff_qty = kwargs.get("roundoff_qty", True)
        new_rows = []

        cgst_sgst_per = float(row[cls.column_to_idx["SGST %"]])
        igst_per = float(row[cls.column_to_idx["IGST %"]])
        gst_percentage = int(cgst_sgst_per * 2)
        amount = float(row.iloc[cls.column_to_idx["Amount"]])

        if amount == 0.0:
            return []

        cgst_sgst_amt = round(amount * cgst_sgst_per / 100, 2)
        igst_amt = round(amount * igst_per / 100, 2)
        discount = 0.0
        total_amt = round(amount + (2 * cgst_sgst_amt) + igst_amt - discount, 2)
        party_cash = (
            "SWIP CARD"
            if cls.is_present_in_mapping(value_to_find=cls.get(row, "Inv No."))
            else cls.get(row, "Party/Cash")
        )

        qty = math.ceil(row[cls.column_to_idx["Qty"]]) if roundoff_qty else row[cls.column_to_idx["Qty"]]
        new_row = cls.get_default_row_format(
            {
                **{
                    tg: str(row.iloc[cls.column_to_idx[tg]]).strip()
                    for tg in cls.target_columns_index
                },
                "Bill No.": bill_no,
                "Party/Cash": party_cash,
                "Qty": qty,
                "Reg Type": "unregistered/consumer",
                "Product's Name": f"Medicine {gst_percentage}%",
                "GST%": gst_percentage,
                "Amount": amount,
                "CGST %": cgst_sgst_per,
                "CGST Amt": cgst_sgst_amt,
                "SGST %": cgst_sgst_per,
                "SGST Amt": cgst_sgst_amt,
                "IGST Amt": igst_amt,
                "Discount %": discount,
                "Total": total_amt,
            }
        )
        new_rows.append(copy.deepcopy(new_row))

        return new_rows

    def transform(cls, path: str, save=True, *args, **kwargs):
        df = pd.read_excel(path)
        # drop the 1st empty column
        # drop the 1st empty column
        df = df.iloc[:, 1:]

        cls.transform_mapping()

        rows_to_df = []
        bill_counter = cls.bill_no_suffix_counter - 1
        last_row = None

        # Iterate over rows
        for _, row in df.iterrows():
            # if gross total reached then break the loop

            if str(row.iloc[1]).strip() == "Gross Total":
                break

            # if cell-0 is nan and we have last row data
            if str(row.iloc[0]).strip() == "nan" and last_row is not None:
                for copy_idx in [1, 2, 3, 4, 5, 6, 7]:
                    row.iloc[copy_idx] = last_row.iloc[copy_idx]

            # if the row does not match the following format= "<digit><dot>"
            elif not str(row.iloc[0]).split(".")[0].strip().isdigit():
                continue
            else:
                # digit case found
                bill_counter += 1
                last_row = copy.deepcopy(row)

            new_tranformed_rows = cls.get_transformed_rows(
                row,
                bill_no=f"{cls.bill_no_prefix}{bill_counter:05d}",
                *args,
                **kwargs
            )

            # if new rows added then only increase the bill counter
            # there may be case where no rows were added because amount is 0
            if len(new_tranformed_rows) > 0:
                rows_to_df += new_tranformed_rows

        result_df = pd.DataFrame(rows_to_df, columns=cls.output_columns)

        return cls.post_processing(
            result_df, save=save, name_prefix=cls.xl_name_prefix, re_arrange_cols=True
        )


class TransformExcelJJOnly(BaseTransformExcel):
    NAME = "JJ/Only"

    def __init__(cls, config: dict, *args, **kwargs) -> None:
        # Product's Name: "Article Description" + " " + "EAN Number"
        # Amount : Rate * Qty
        # Invoice No: <prefix><counter>

        cls.output_columns = config["output_columns"]
        cls._default_output_row = config["default_output_row"]
        cls.columns_to_sum = config["columns_to_sum"]
        cls.target_columns_index = config["target_columns_index"]
        cls.file_save_dir = "transformed"
        cls.perfix_for_totals = config["perfix_for_totals"]
        cls.column_to_idx: dict = config["column_to_idx"]

        # changes every time
        cls.bill_no_prefix = kwargs.get("bill_no_prefix", None)
        cls.bill_no_suffix_counter = kwargs.get("bill_no_suffix_counter", None)

    def get_transformed_rows(cls, row, inv_no: str, calculate_igst=False):
        new_rows = []
        gst_percentage = int(cls.get(row, "GST %"))
        rate = round(float(cls.get(row, "Rate")), 2)
        qty = math.ceil(cls.get(row, "Qty"))
        amount = round(rate * qty, 2)
        discount = 0.0

        cgst_sgst_per = gst_percentage / 2

        if amount == 0.0:
            return None

        cgst_sgst_amt = round(amount * cgst_sgst_per / 100, 2)
        igst_amt = round(amount * gst_percentage / 100, 2) if calculate_igst else 0.0
        total_amt = round(amount + (2 * cgst_sgst_amt) + igst_amt - discount, 2)
        inv_date = cls.get(row, "Inv Date").strftime(format="%d/%m/%Y")

        # party/cash
        party_cash = (
            cls.get(row, "Article Description").strip()
            + " "
            + cls.get(row, "EAN Number").strip()
        )

        new_row = cls.get_default_row_format(
            {
                **{
                    colname: str(cls.get(row, colname)).strip()
                    for colname in cls.target_columns_index
                },
                "Inv Date": inv_date,
                "Inv No.": inv_no,
                "Product's Name": party_cash,
                "GST %": gst_percentage,
                "Amount": amount,
                "CGST %": cgst_sgst_per,
                "CGST Amt": cgst_sgst_amt,
                "SGST %": cgst_sgst_per,
                "SGST Amt": cgst_sgst_amt,
                "IGST Amt": igst_amt,
                "Discount %": discount,
                "Total": total_amt,
            }
        )
        new_rows.append(copy.deepcopy(new_row))

        return new_rows

    def transform(cls, path: str, save=True):
        df = pd.read_excel(path)
        rows_to_df = []
        bill_counter = cls.bill_no_suffix_counter

        # Removing the first 2 rows and first col
        df = df.iloc[2:, 1:]
        df.reset_index(drop=True, inplace=True)

        billNo_counter_dict = dict()

        # Iterate over rows
        for _, row in df.iterrows():
            row = list(row)

            if billNo_counter_dict.get(cls.get(row, "Bill No.")) is not None:
                _bill_counter = billNo_counter_dict.get(cls.get(row, "Bill No."))
            else:
                _bill_counter = bill_counter
                # add the new bill number to dict
                billNo_counter_dict[cls.get(row, "Bill No.")] = _bill_counter
                bill_counter += 1  # update the bill no

            # Set current bill number
            new_tranformed_rows = cls.get_transformed_rows(
                row,
                inv_no=f"{cls.bill_no_prefix}{_bill_counter:05d}",
            )
            rows_to_df += new_tranformed_rows

        # File saving starts from here
        filename = cls.get_filename_with_datetime(prefix="JJOnly_")
        xl_save_path = os.path.join(cls.file_save_dir, filename)

        # if dir not exist then create one
        if not os.path.isdir(cls.file_save_dir):
            os.makedirs(cls.file_save_dir, exist_ok=True)

        writer = pd.ExcelWriter(xl_save_path, engine="xlsxwriter")

        result_df = pd.DataFrame(rows_to_df, columns=cls.output_columns)

        # Convert the dataframe to an XlsxWriter Excel object.
        result_df.to_excel(writer, sheet_name="Sheet1", index=False)

        # Get the xlsxwriter objects from the dataframe writer object.
        workbook = writer.book
        worksheet = writer.sheets["Sheet1"]

        no_of_row = worksheet.dim_rowmax
        for col_to_sum in cls.columns_to_sum:
            col_letter = get_column_letter(
                result_df.columns.get_loc(col_to_sum) + 1
            )  # offset of 1 for index to pos

            index_of_sheet = f"{col_letter}{no_of_row + 2}"
            formula = f"=SUM({col_letter}2:{col_letter}{no_of_row + 1})"

            worksheet.write_formula(index_of_sheet, formula)

        if len(cls.columns_to_sum) > 0:
            col_letter = get_column_letter(
                result_df.columns.get_loc(cls.perfix_for_totals["column"]) + 1
            )  # offset of 1 for index to pos
            index_of_sheet = f"{col_letter}{no_of_row + 2}"
            worksheet.write_string(index_of_sheet, cls.perfix_for_totals["label"])

        if save:
            print(f"File saved to: {xl_save_path}")
            writer.close()

        return dict(
            xl_save_path=xl_save_path, save_dir=cls.file_save_dir, xl_file_name=filename
        )


class GSTR1Self(BaseTransformExcel):
    NAME = "GSTR1 Self"
    CONFIG = "gstr1_self_config.json"

    def __init__(cls, config, **kwargs):
        cls.output_columns = config["output_columns"]
        cls._default_output_row = config["default_output_row"]
        cls.columns_to_sum = config["columns_to_sum"]
        cls.direct_target_cols = config["direct_target_cols"]
        cls.perfix_for_totals = config["perfix_for_totals"]
        cls.gst_data: dict = config["gst_data"]

        # kwargs
        cls.calculate_igst = kwargs.get("calculate_igst", False)
        cls.inv_no_prefix = kwargs["inv_no_prefix"]
        cls.inv_no_suffix_counter = kwargs["inv_no_suffix_counter"]

        # constants
        cls.file_save_dir = "transformed"
        cls.filename_prefix = "GSTR1_Self_"

    def remove_first_digits(cls, inp: str):
        idx = 0

        for c in inp:
            if not c.isdigit():
                break
            idx += 1

        return (inp[idx:]).strip()

    def get_transformed_rows(cls, row, inv_no: str, calculate_igst=False) -> list:
        new_rows = []

        for gst_data in cls.gst_data:
            gst_percentage = gst_data["gst_percentage"]
            cgst_sgst_per = gst_percentage / 2

            amount = round(float(row["RATE"]) * float(row["QTY"]), 2)
            cgst_sgst_amt = round(amount * cgst_sgst_per / 100, 2)
            igst_amt = (
                round(amount * gst_percentage / 100, 2) if calculate_igst else 0.0
            )
            total_amt = round(amount + (2 * cgst_sgst_amt) + igst_amt, 2)
            narration = f"Bill No. {str(row['BILL NO']).strip()} {str(row['PARTY']).strip()} {str(row['ITEM NAME']).strip()}"

            new_row = cls.get_default_row_format(
                {
                    **{
                        i["new_col"]: str(row[i["from_col"]]).strip()
                        for i in cls.direct_target_cols
                    },
                    "Inv No.": inv_no,
                    "GST%": gst_percentage,
                    "Party/Cash": cls.remove_first_digits(row["PARTY"]),
                    "Amount": amount,
                    "CGST %": cgst_sgst_per,
                    "CGST Amt": cgst_sgst_amt,
                    "SGST %": cgst_sgst_per,
                    "SGST Amt": cgst_sgst_amt,
                    "IGST Amt": igst_amt,
                    "Total": total_amt,
                    "Narration": narration,
                }
            )
            new_rows.append(copy.deepcopy(new_row))
        return new_rows

    def transform(cls, path: str, save=True):
        # skip 6 rows
        df = pd.read_excel(path, skiprows=6)
        rows_to_df = []

        inv_no_getter = dict()
        inv_counter = cls.inv_no_suffix_counter

        for _, row in df.iterrows():
            # guard clause for amt == 0
            try:
                if float(row["RATE"]) * float(row["QTY"]) == 0.0:
                    continue
                if str(row["BILL NO"]).strip() == "nan":
                    continue
            except Exception as e:
                continue

            # set the invoice number for bill no
            if inv_no_getter.get(str(row["BILL NO"]).strip()) is None:
                inv_no_getter[str(row["BILL NO"]).strip()] = inv_counter
                inv_counter += 1

            inv_suffix = inv_no_getter.get(str(row["BILL NO"]).strip())
            rows_to_df += cls.get_transformed_rows(
                row,
                inv_no=f"{cls.inv_no_prefix}{inv_suffix}",
                calculate_igst=cls.calculate_igst,
            )

        filename = cls.get_filename_with_datetime(prefix=cls.filename_prefix)
        xl_save_path = os.path.join(cls.file_save_dir, filename)

        # if dir not exist then create one
        if not os.path.isdir(cls.file_save_dir):
            os.makedirs(cls.file_save_dir, exist_ok=True)

        writer = pd.ExcelWriter(xl_save_path, engine="xlsxwriter")

        result_df = pd.DataFrame(rows_to_df, columns=cls.output_columns)

        # Convert the dataframe to an XlsxWriter Excel object.
        result_df.to_excel(writer, sheet_name="Sheet1", index=False)

        # Get the xlsxwriter objects from the dataframe writer object.
        workbook = writer.book
        worksheet = writer.sheets["Sheet1"]

        no_of_row = worksheet.dim_rowmax
        for col_to_sum in cls.columns_to_sum:
            col_letter = get_column_letter(
                result_df.columns.get_loc(col_to_sum) + 1
            )  # offset of 1 for index to pos

            index_of_sheet = f"{col_letter}{no_of_row + 2}"
            formula = f"=SUM({col_letter}2:{col_letter}{no_of_row + 1})"

            worksheet.write_formula(index_of_sheet, formula)

        if len(cls.columns_to_sum) > 0:
            col_letter = get_column_letter(
                result_df.columns.get_loc(cls.perfix_for_totals["column"]) + 1
            )  # offset of 1 for index to pos
            index_of_sheet = f"{col_letter}{no_of_row + 2}"
            worksheet.write_string(index_of_sheet, cls.perfix_for_totals["label"])

        if save:
            print(f"File saved to: {xl_save_path}")
            writer.close()

        return dict(
            xl_save_path=xl_save_path, save_dir=cls.file_save_dir, xl_file_name=filename
        )


class SMCGlobalShare(BaseTransformExcel):
    APP_NAME = "SMC GLOBAL / SHARE"

    def __init__(cls) -> None:
        cls.filename_prefix = "SMC-GLOBAL-SHARE_"
        cls.file_save_dir = "transformed"

    def transform(cls, path: str, save=True):
        # skip 6 rows
        df = pd.read_excel(path, skiprows=13)
        rows_to_df = []

        df = df[~df["Code"].str.startswith("Total Of(")]

        filename = cls.get_filename_with_datetime(prefix=cls.filename_prefix)
        xl_save_path = os.path.join(cls.file_save_dir, filename)

        # if dir not exist then create one
        if not os.path.isdir(cls.file_save_dir):
            os.makedirs(cls.file_save_dir, exist_ok=True)

        if save:
            df.to_excel(xl_save_path, sheet_name="Sheet1", index=False)
            print(f"File saved to: {xl_save_path}")

        return dict(
            xl_save_path=xl_save_path, save_dir=cls.file_save_dir, xl_file_name=filename
        )


class GSTR1EchsPmjay(BaseTransformExcel):
    APP_NAME = "GSTR1 ECHS / PMJAY"  # This will send to the display.html
    CONFIG = "gstr1_echs_pmjay_config.json"

    def __init__(cls, config: dict, *args, **kwargs) -> None:
        cls.file_save_dir = "transformed"
        cls.output_columns = config["output_columns"]
        cls._default_output_row = config["default_output_row"]
        cls.columns_to_sum = config["columns_to_sum"]
        cls.perfix_for_totals = config["perfix_for_totals"]
        cls.column_to_idx: dict = config["column_to_idx"]
        cls.target_columns_index: list = config["target_columns_index"]

        # changes every time
        cls.bill_no_prefix = kwargs.get("bill_no_prefix", None)
        cls.bill_no_suffix_counter = kwargs.get("bill_no_suffix_counter", None)
        cls.calculate_igst = kwargs.get("calculate_igst", False)

        # Mapping file
        cls._mapping_df = kwargs.get("mapping_df", None)  # optional
        cls.mapping_sheetname = config["mapping_sheetname"]

    def transform_mapping(cls):
        if cls._mapping_df is None:
            return

        df = cls._mapping_df.iloc[6:, 1:]
        # Set the headers to be the values from the first row
        new_headers = df.iloc[0]
        df.columns = new_headers

        # Reset the index
        df.reset_index(drop=True, inplace=True)

        cls._mapping_df = df

    def is_present_in_mapping(cls, row):
        if cls._mapping_df is None:
            return cls.get(row, "Party/Cash")

        inv_no = cls.get(row, "Inv No.")

        if (
            inv_no is None or str(inv_no).lower() == "nan"
        ):  # if inv no not found in mapping return original party/cash
            return cls.get(row, "Party/Cash")

        if inv_no not in cls._mapping_df["Bill No."].values:
            return cls.get(row, "Party/Cash")

        filtered_row = cls._mapping_df[cls._mapping_df["Bill No."] == inv_no]
        card_payment = filtered_row["CARD PAYMENT"]

        if str(card_payment).lower() == "nan":
            return cls.get(row, "Party/Cash")

        if float(card_payment) > 0.0:
            return "SWIP CARD"

        if float(card_payment) == 0.0:
            if str(filtered_row["Name of the Person"]).lower() in ["echs", "pmjay"]:
                return filtered_row["Name of the Person"]
            else:
                # Take the original party name from row
                return cls.get(row, "Party/Cash")

        return f"NegValue<{inv_no}>"

    def get_transformed_rows(cls, row, bill_no: str):
        new_rows = []

        cgst_sgst_per = float(row[cls.column_to_idx["SGST %"]])
        igst_per = float(row[cls.column_to_idx["IGST %"]])
        gst_percentage = int(cgst_sgst_per * 2)
        amount = float(row.iloc[cls.column_to_idx["Amount"]])

        if amount == 0.0:
            return []

        cgst_sgst_amt = round(amount * cgst_sgst_per / 100, 2)
        igst_amt = round(amount * igst_per / 100, 2)
        discount = 0.0
        total_amt = round(amount + (2 * cgst_sgst_amt) + igst_amt - discount, 2)
        party_cash = cls.is_present_in_mapping(row)

        new_row = cls.get_default_row_format(
            {
                **{
                    tg: str(row.iloc[cls.column_to_idx[tg]]).strip()
                    for tg in cls.target_columns_index
                },
                "Bill No.": bill_no,
                "Party/Cash": party_cash,
                "Qty": math.ceil(row[cls.column_to_idx["Qty"]]),
                "Reg Type": "unregistered/consumer",
                "Product's Name": f"Medicine {gst_percentage}%",
                "GST%": gst_percentage,
                "Amount": amount,
                "CGST %": cgst_sgst_per,
                "CGST Amt": cgst_sgst_amt,
                "SGST %": cgst_sgst_per,
                "SGST Amt": cgst_sgst_amt,
                "IGST Amt": igst_amt,
                "Discount %": discount,
                "Total": total_amt,
            }
        )
        new_rows.append(copy.deepcopy(new_row))

        return new_rows

    def transform(cls, path: str, save=True):
        df = pd.read_excel(path)
        # drop the 1st empty column
        # drop the 1st empty column
        df = df.iloc[:, 1:]

        cls.transform_mapping()

        rows_to_df = []
        bill_counter = cls.bill_no_suffix_counter - 1
        last_row = None

        # Iterate over rows
        for _, row in df.iterrows():
            # if gross total reached then break the loop

            if str(row.iloc[1]).strip() == "Gross Total":
                break

            # if cell-0 is nan and we have last row data
            if str(row.iloc[0]).strip() == "nan" and last_row is not None:
                for copy_idx in [1, 2, 3, 4, 5, 6, 7]:
                    row.iloc[copy_idx] = last_row.iloc[copy_idx]

            # if the row does not match the following format= "<digit><dot>"
            elif not str(row.iloc[0]).split(".")[0].strip().isdigit():
                continue
            else:
                # digit case found
                bill_counter += 1
                last_row = copy.deepcopy(row)

            new_tranformed_rows = cls.get_transformed_rows(
                row,
                bill_no=f"{cls.bill_no_prefix}{bill_counter:05d}",
            )

            # if new rows added then only increase the bill counter
            # there may be case where no rows were added because amount is 0
            if len(new_tranformed_rows) > 0:
                rows_to_df += new_tranformed_rows

        filename = cls.get_filename_with_datetime(prefix="GSTR1EchsPmjay_")
        xl_save_path = os.path.join(cls.file_save_dir, filename)

        # if dir not exist then create one
        if not os.path.isdir(cls.file_save_dir):
            os.makedirs(cls.file_save_dir, exist_ok=True)

        writer = pd.ExcelWriter(xl_save_path, engine="xlsxwriter")

        result_df = pd.DataFrame(rows_to_df, columns=cls.output_columns)

        # Convert the dataframe to an XlsxWriter Excel object.
        result_df.to_excel(writer, sheet_name="Sheet1", index=False)

        # Get the xlsxwriter objects from the dataframe writer object.
        worksheet = writer.sheets["Sheet1"]

        no_of_row = worksheet.dim_rowmax
        for col_to_sum in cls.columns_to_sum:
            col_letter = get_column_letter(
                result_df.columns.get_loc(col_to_sum) + 1
            )  # offset of 1 for index to pos

            index_of_sheet = f"{col_letter}{no_of_row + 2}"
            formula = f"=SUM({col_letter}2:{col_letter}{no_of_row + 1})"

            worksheet.write_formula(index_of_sheet, formula)

        if len(cls.columns_to_sum) > 0:
            col_letter = get_column_letter(
                result_df.columns.get_loc(cls.perfix_for_totals["column"]) + 1
            )  # offset of 1 for index to pos
            index_of_sheet = f"{col_letter}{no_of_row + 2}"
            worksheet.write_string(index_of_sheet, cls.perfix_for_totals["label"])

        if save:
            print(f"File saved to: {xl_save_path}")
            writer.close()

        return dict(
            xl_save_path=xl_save_path, save_dir=cls.file_save_dir, xl_file_name=filename
        )


class GSTR1Marg(BaseTransformExcel):
    APP_NAME = "GSTR1 Marg"  # This will send to the display.html
    CONFIG = "gstr1_marg_config.json"

    def __init__(cls, config: dict, *args, **kwargs) -> None:
        cls.file_save_dir = "transformed"
        cls.output_columns = config["output_columns"]
        cls._default_output_row = config["default_output_row"]
        cls.columns_to_sum = config["columns_to_sum"]
        cls.perfix_for_totals = config["perfix_for_totals"]
        cls.column_to_idx: dict = config["column_to_idx"]
        cls.target_columns_index: list = config["target_columns_index"]

        # changes every time
        cls.bill_no_prefix = kwargs.get("bill_no_prefix", None)
        cls.bill_no_suffix_counter = kwargs.get("bill_no_suffix_counter", None)
        cls.calculate_igst = kwargs.get("calculate_igst", False)

        # Mapping file
        cls._mapping_df = kwargs.get("mapping_df", None)  # optional
        cls.mapping_sheetname = config["mapping_sheetname"]

    def transform_mapping(cls):
        if cls._mapping_df is None:
            return

        df = cls._mapping_df.iloc[6:, 1:]
        # Set the headers to be the values from the first row
        new_headers = df.iloc[0]
        df.columns = new_headers

        # Reset the index
        df.reset_index(drop=True, inplace=True)

        cls._mapping_df = df

    def is_present_in_mapping(cls, row):
        if cls._mapping_df is None:
            return "CASH"

        inv_no = cls.get(row, "Inv No.")

        if (
            inv_no is None or str(inv_no).lower() == "nan"
        ):  # if inv no not found in mapping return original party/cash
            return "CASH"

        if inv_no not in cls._mapping_df["Bill No."].values:
            return "CASH"

        filtered_row = cls._mapping_df[cls._mapping_df["Bill No."] == inv_no]
        card_payment = filtered_row["CARD PAYMENT"]

        if str(card_payment).lower() == "nan":
            return "CASH"

        if float(card_payment) > 0.0:
            return "SWIP CARD"

        if float(card_payment) == 0.0:
            if str(filtered_row["Name of the Person"]).lower() in ["echs", "pmjay"]:
                return filtered_row["Name of the Person"]
            else:
                # Take the original party name from row
                return "CASH"

        return f"NegValue<{card_payment}>"

    def get_transformed_rows(cls, row, bill_no: str, row_data_dict: dict):

        # amount = float(row.iloc[cls.column_to_idx["Amount"]])
        taxable_amt = float(row.iloc[cls.column_to_idx["Taxable Amount"]])

        if taxable_amt == 0.0:
            return None  # if taxable_amt if zero for any bill then return none

        # set initial values for bill no
        if row_data_dict.get(bill_no) is None:

            gst_no = cls.is_present_in_mapping(row)
            _row_data = cls.get_default_row_format(
                {
                    **{
                        tg: str(row.iloc[cls.column_to_idx[tg]]).strip()
                        for tg in cls.target_columns_index
                    },
                    "GST.No.": gst_no,
                    "PARTY NAME": cls.get(row, "Party/Cash"),
                    "BILL NO.": str(bill_no),
                }
            )

        else:
            _row_data = row_data_dict[bill_no]

        cgst_sgst_per = float(row[cls.column_to_idx["SGST %"]])
        gst_percentage = int(cgst_sgst_per * 2)
        cgst_sgst_amt = round(taxable_amt * cgst_sgst_per / 100, 2)

        _row_data[f"CGST {cgst_sgst_per}%"] += cgst_sgst_amt
        _row_data[f"SGST {cgst_sgst_per}%"] += cgst_sgst_amt
        _row_data[f"GST {gst_percentage}%"] += taxable_amt

        def round_and_diff(num):
            rounded_num = round(num)
            difference = rounded_num - num
            return round(difference, 2)

        def get_totalAmt_nd_roundOff():
            amt = (
                _row_data["GST 5%"]
                + _row_data["CGST 2.5%"]
                + _row_data["SGST 2.5%"]
                + _row_data["GST 12%"]
                + _row_data["CGST 6.0%"]
                + _row_data["SGST 6.0%"]
                + _row_data["GST 18%"]
                + _row_data["CGST 9.0%"]
                + _row_data["CGST 9.0%"]
            )

            return round(amt), round_and_diff(amt)

        _row_data[f"BILL VALUE"], _row_data[f"Roundoff"] = get_totalAmt_nd_roundOff()

        return copy.deepcopy(_row_data)

    def transform(cls, path: str, save=True):
        df = pd.read_excel(path)
        # drop the 1st empty column
        # drop the 1st empty column
        df = df.iloc[:, 1:]

        cls.transform_mapping()

        rows_to_df = []
        bill_counter = cls.bill_no_suffix_counter - 1
        last_row = None

        row_data_dict = dict()  # bill no: data
        inv_no_list = []

        # Iterate over rows
        for _, row in df.iterrows():
            # row = row.tolist()
            # if gross total reached then break the loop

            if str(row.iloc[1]).strip() == "Gross Total":
                break

            # if cell[0] is nan and we have last row data
            if str(row.iloc[0]).strip() == "nan" and last_row is not None:
                for copy_idx in [1, 2, 3, 4, 5, 6, 7]:
                    row.iloc[copy_idx] = last_row.iloc[copy_idx]

            # if the row does not match the following format= "<digit><dot>"
            elif not str(row.iloc[0]).split(".")[0].strip().isdigit():
                continue
            else:
                # digit case found
                bill_counter += 1
                last_row = copy.deepcopy(row)

            bill_no = f"{cls.bill_no_prefix}{bill_counter:05d}"

            row_data_dict[bill_no] = cls.get_transformed_rows(
                row,
                bill_no=bill_no,
                row_data_dict=row_data_dict,
            )

            if row_data_dict[bill_no] is None:
                del row_data_dict[bill_no]

        rows_to_df = list(row_data_dict.values())
        # rows_to_df = [value for value in row_data_dict.values() if value is not None]

        result_df = pd.DataFrame(rows_to_df, columns=cls.output_columns)

        # print(row_data_dict.keys())
        filename = cls.get_filename_with_datetime(prefix="GSTR1Marg_")
        xl_save_path = os.path.join(cls.file_save_dir, filename)

        # if dir not exist then create one
        if not os.path.isdir(cls.file_save_dir):
            os.makedirs(cls.file_save_dir, exist_ok=True)

        writer = pd.ExcelWriter(xl_save_path, engine="xlsxwriter")

        # Convert the dataframe to an XlsxWriter Excel object.
        result_df.to_excel(writer, sheet_name="Sheet1", index=False)

        # Get the xlsxwriter objects from the dataframe writer object.
        worksheet = writer.sheets["Sheet1"]

        no_of_row = worksheet.dim_rowmax
        for col_to_sum in cls.columns_to_sum:
            col_letter = get_column_letter(
                result_df.columns.get_loc(col_to_sum) + 1
            )  # offset of 1 for index to pos

            index_of_sheet = f"{col_letter}{no_of_row + 2}"
            formula = f"=SUM({col_letter}2:{col_letter}{no_of_row + 1})"

            worksheet.write_formula(index_of_sheet, formula)

        if len(cls.columns_to_sum) > 0:
            col_letter = get_column_letter(
                result_df.columns.get_loc(cls.perfix_for_totals["column"]) + 1
            )  # offset of 1 for index to pos
            index_of_sheet = f"{col_letter}{no_of_row + 2}"
            worksheet.write_string(index_of_sheet, cls.perfix_for_totals["label"])

        if save:
            print(f"File saved to: {xl_save_path}")
            writer.close()

        return dict(
            xl_save_path=xl_save_path, save_dir=cls.file_save_dir, xl_file_name=filename
        )


class GSTR1WQty_Extended(BaseTransformExcel):
    HTML_ID = "transform-gstr1_w_qty_sum"
    APP_NAME = "GSTR1 With Qty (sum)"  # This will send to the display.html
    CONFIG = "gstWithQtySum_config.json"

    def __init__(cls, config: dict, *args, **kwargs) -> None:
        cls.file_save_dir = "transformed"

        cls.output_columns = config["output_columns"]
        cls._default_output_row = config["default_output_row"]
        cls.columns_to_sum = config["columns_to_sum"]
        cls.perfix_for_totals = config["perfix_for_totals"]
        cls.column_to_idx: dict = config["column_to_idx"]
        cls.target_columns_index: list = config["target_columns_index"]
        cls.xl_name_prefix: str = config["xl_name_prefix"]

        # changes every time
        cls.bill_no_prefix = kwargs.get("bill_no_prefix", None)
        cls.bill_no_suffix_counter = kwargs.get("bill_no_suffix_counter", None)
        cls.calculate_igst = kwargs.get("calculate_igst", False)

        # pre transformation
        cls._gst_w_qty_transformation = TransformExcelGST(config, *args, **kwargs)

    def preprocess_df(cls, df: Union[str, pd.DataFrame]) -> pd.DataFrame:
        return cls._gst_w_qty_transformation.transform(df, save=False, roundoff_qty=False)["df"]

    def transform(cls, df: str, save=True):

        df = cls.preprocess_df(df)
        print(f"After preprocessing: {df.shape}")

        # sort the values
        df = df.sort_values(
            # by=["Inv Date", "Product's Name", "Party/Cash"], ascending=False
            by=["Inv Date", "Product's Name", "Party/Cash"], ascending=True
        )

        df = df if isinstance(df, pd.DataFrame) else pd.read_excel(df)

        # Step 1: Filter distinct values for "Inv Date" column
        distinct_dates = df["Inv Date"].unique()

        # Initialize an empty DataFrame to store the final result
        result_df = pd.DataFrame()
        bill_counter = cls.bill_no_suffix_counter
        prv_bill_no = None

        # Iterate over distinct dates
        for date in distinct_dates:
            # Filter DataFrame for the current date
            filtered_df = df[df["Inv Date"] == date]

            map_BillNo_to_partyCash = {
                "cash": None,
                "echs": None,
                "pmjay": None,
                "swip card": None,
                "shubhanu eye hospital": None,
            }

            # Step 4: Filter distinct values (eg Medicine 5%, Medicine 18%) for "Product's Name" column
            distinct_products = filtered_df["Product's Name"].unique()
            # Iterate over distinct products
            for product in distinct_products:
                # Filter DataFrame for the current product
                filtered_df2 = filtered_df[filtered_df["Product's Name"] == product]

                # Change the normal party name to cash
                filtered_df2["Party/Cash Copy"] = filtered_df2["Party/Cash"]
                filtered_df2["Party/Cash"] = filtered_df2["Party/Cash"].apply(
                    lambda x: (
                        x
                        if str(x).lower().strip()
                        in ["echs", "pmjay", "swip card", "shubhanu eye hospital"]
                        else "CASH"
                    )
                )

                # Step 2: Filter distinct values for "Party/Cash" column
                distinct_parties = filtered_df2["Party/Cash"].unique()

                # Iterate over distinct parties
                for party in distinct_parties:
                    filtered_df3 = filtered_df2[filtered_df2["Party/Cash"] == party]

                    # Sum the int and float values
                    sums = (
                        # filtered_df3.select_dtypes(include=["int", "float"])
                        filtered_df3[[
                            "Qty", "Amount", "CGST Amt", "SGST Amt", "IGST Amt", "Total"
                        ]]
                        .sum()
                        .to_dict()
                    )
                    # print("========================================================")
                    # print(f"{date=} {product=} {party=}")
                    # print(sums)


                    # Join all elements to form a single string
                    narration = ", ".join(
                        "("
                        + filtered_df3["Inv No."]
                        + ", "
                        + filtered_df3["Party/Cash Copy"]
                        + ")"
                    )

                    _temp_bill_counter = map_BillNo_to_partyCash.get(party)
                    if _temp_bill_counter is None:
                        _temp_bill_counter = bill_counter
                        map_BillNo_to_partyCash[party] = _temp_bill_counter
                        bill_counter += 1

                    # Step 5: Create a DataFrame from sums and strings dictionaries
                    df_sums = pd.DataFrame(
                        {
                            **sums,
                            "Bill No.": f"{cls.bill_no_prefix}{_temp_bill_counter :05d}",
                            "Narration": narration,
                            **{
                                s: filtered_df3[s].iloc[0]
                                for s in [
                                    "GST%",	
                                    "CGST %",
                                    "SGST %",
                                    "Rate",
                                    "Discount %",
                                    "Party/Cash",
                                    "Product's Name",
                                    "State",
                                    "Party's GST",
                                    "Inv No.",
                                    "Inv Date",
                                    "HSN Code",
                                    "Reg Type",
                                    "Place of Supply",
                                    "Country",
                                    "Consignee State",
                                    "Consignee GST",
                                ]
                            },
                        },
                        index=[0],
                    )

                    # # increment ! when prv_bill_no does not matches current bill no
                    # if prv_bill_no !=  str(filtered_df3['Bill No.'].iloc[0]):
                    #     # increment bill no
                    #     bill_counter += 1

                    prv_bill_no = filtered_df3["Bill No."].iloc[0]

                    # Append the DataFrame to result_df
                    # result_df = result_df.append(df_sums, ignore_index=True)
                    result_df = pd.concat([result_df, df_sums], ignore_index=True)

        # sort the values
        result_df = result_df.sort_values(by=["Bill No."], ascending=True)

        res = cls.post_processing(
            df=result_df,
            save=save,
            name_prefix=cls.xl_name_prefix,
            re_arrange_cols=True,
        )

        print(f"After preprocessing: {res['df'].shape}")

        return res
        # if save:
        #     result_df.to_excel(cls.xl_name_prefix, index=False)
        # return result_df


## Sanity check functions
def check_for_stock():
    default_config = {}
    obj = TransformStockExcel(default_config)
    obj.transform(
        "/home/akshat/Documents/projects/joshi-uncle/data/MutalFund_stock_forErrorRows.xlsx",
        save=True,
    )


def check_for_purchase():
    default_config = TransformExcelPurchase.default_config()

    obj = TransformExcelPurchase(default_config)

    def get_df():
        sale_path = "/home/akshat/Documents/projects/joshi-uncle/data/REPORT_2222.XLS"
        df = pd.read_excel(
            sale_path, sheet_name="MARG ERP 9+ Excel Report", header=None
        )
        return df

    obj.transform_bill(get_df())


def check_for_sale():
    default_config = TransformExcelSale.default_config()

    obj = TransformExcelSale(default_config)

    def get_df():
        sale_path = "data/SALE JAN FINAL 2024.XLS"
        df = pd.read_excel(
            sale_path, sheet_name="MARG ERP 9+ Excel Report", header=None
        )
        return df

    obj.transform_bill(get_df())


def check_for_gst():
    default_config = TransformExcelGST.read_config(
        os.path.join("excel_app", "config", "gst_config.json")
    )
    bill_no_prefix = "JPS/23/24/"
    bill_no_suffix_counter = 1546

    obj = TransformExcelGST(
        default_config,
        bill_no_prefix=bill_no_prefix,
        bill_no_suffix_counter=bill_no_suffix_counter,
    )

    def get_df():
        sale_path = "../data/gst2.xls"
        df = pd.read_excel(sale_path, header=None)
        return df

    obj.transform(get_df())


def check_for_jjonly():

    default_config = TransformExcelSale.read_config(
        os.path.join("excel_app", "config", "jj-only_config.json")
    )
    bill_no_prefix = "TDR/23-24/"
    bill_no_suffix_counter = 1701

    obj = TransformExcelJJOnly(
        default_config,
        bill_no_prefix=bill_no_prefix,
        bill_no_suffix_counter=bill_no_suffix_counter,
    )

    def get_df():
        sale_path = "../data/jj-only.xlsx"
        df = pd.read_excel(sale_path, header=None)
        return df

    obj.transform(get_df())


def check_for_gst2b():
    default_config = TransformExcelSale.read_config(
        os.path.join("excel_app", "config", "gst2b_config.json")
    )
    obj = GST2BTransfromation(default_config)

    def get_df():
        dfs = []
        for file in os.listdir("../data/GST2BExcels/"):
            dfs.append("../data/GST2BExcels/" + file)

        return dfs

    dfs = get_df()
    obj.transform(dfs)


def check_for_echs():
    default_config = TransformExcelSale.read_config(
        os.path.join("excel_app", "config", "echs_due-config.json")
    )
    obj = EchsDueTransfromation(default_config)

    def get_df():
        dfs = []
        for file in os.listdir("../data/echs-due/"):
            dfs.append("../data/echs-due/" + file)

        return dfs

    dfs = get_df()
    obj.transform(dfs)


def check_for_gstr1Self():
    default_config = TransformExcelSale.read_config(
        os.path.join("excel_app", "config", "gstr1_self_config.json")
    )
    obj = GSTR1Self(default_config)

    obj.transform("../data/GSTR1_self.XLS")


def check_for_SMCGlobalShare():
    obj = SMCGlobalShare()

    obj.transform("../data/SMC-GLOBAL-SHARE.XLSx")


def check_for_GSTR1Marg():
    mapping_df = pd.read_excel(
        "/home/akshat/Documents/projects/joshi-uncle/data/GSTR 1 ECHS -PMJAY/gstr1_echs_pmjay_mapping.xls"
    )

    default_config = GSTR1Marg.read_config(
        os.path.join("excel_app", "config", "gstr1_marg_config.json")
    )
    bill_no_prefix = "JSP/23/23/"
    bill_no_suffix_counter = 123

    excel_file = "/home/akshat/Documents/projects/joshi-uncle/data/GSTR 1 ECHS -PMJAY/gstr1_echs_pmjay.xls"
    transform = GSTR1Marg(
        default_config,
        bill_no_prefix=bill_no_prefix,
        bill_no_suffix_counter=bill_no_suffix_counter,
        mapping_df=mapping_df,
    )
    data = transform.transform(excel_file, save=True)


def check_for_gstWQty_extented():
    default_config = TransformExcelGST.read_config(
        os.path.join("excel_app", "config", "gstWithQtySum_config.json")
    )
    bill_no_prefix = "JPS/23/24/"
    bill_no_suffix_counter = 1546

    obj2 = GSTR1WQty_Extended(
        default_config,
        bill_no_prefix=bill_no_prefix,
        bill_no_suffix_counter=bill_no_suffix_counter,
        mapping_df=pd.read_excel("../data/gstr1_w_qty_mapping.xls"),
    )
    obj2.transform("../data/gstr1_w_qty.xls")


if __name__ == "__main__":
    check_for_gstWQty_extented()
